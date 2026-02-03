from autograder.result_writer.ResultWriter import ResultWriter, ResultObject
from openpyxl import Workbook, load_workbook
import os
import pandas as pd
import json

# Write actual headers
class ExcelResultWriter(ResultWriter):
    def __init__(self, filePath):
        super().__init__(filePath)

    def InitializeWriter(self):
        self.header_written = False

        if os.path.exists(self.filePath) and os.path.getsize(self.filePath) != 0:
            self.wb = load_workbook(self.filePath)
            self.ws = self.wb.active
            self.start_row = self.ws.max_row + 1
            self.header_written = True  # Assume header already present
        else:
            self.wb = Workbook()
            self.ws = self.wb.active
            self.start_row = 1

    def writeMultiIndexHeader(self, columns):
        levels = columns.nlevels
        for level in range(levels):
            self.ws.append([
                col[level] if len(col) > level else "" for col in columns
            ])
        self.start_row += levels
        self.header_written = True
        self.wb.save(self.filePath)

    def writeTestResult(self, resultObject: ResultObject):
        passedSetDict = {}
        resultObject.TestResults = sorted(resultObject.TestResults, key=lambda result: result.TestName)
        for result in resultObject.TestResults:
            passedSetDict[("Tests", result.TestName, "Passed")] = result.Passed
            passedSetDict[("Tests", result.TestName, "Output")] = result.Output
            passedSetDict[("Tests", result.TestName, "Error")] = result.Error
            # passedSetDict[("Tests", result.TestName, "Similarity Report")] = result.SimilarityReport

        res = {
            ("Student Name","", "") : resultObject.StudentName,
            ("Asu Id","", ""): resultObject.AsuId,
            ("Fatal Errors","", ""): resultObject.FatalErrors,
            ("Build Passed","", ""): resultObject.BuildPassed,
            ("Build Errors","", ""): resultObject.BuildError,
            #("Structure Errors", "", ""): json.dumps(resultObject.StructureErrors),
            #("Submission Error", "", ""): resultObject.SubmissionError,
            ("Final Score","",""): resultObject.FinalScore,
        }
        res.update(passedSetDict)
        self.writeRow(res)
    
    def writeRow(self, row_data):
        if not self.header_written:
            df = pd.DataFrame([row_data])
            df_columns = pd.MultiIndex.from_tuples(df.columns)
            self.writeMultiIndexHeader(df_columns)
        self.ws.append(list(row_data.values()))
        self.wb.save(self.filePath)

    def FinalizeWriter(self):
        pass
